import os, requests

from openpyxl import Workbook
from openpyxl import load_workbook

from datetime import date
from django.conf import settings

from django_cron import CronJobBase, Schedule
from django.db.models import F
from django.utils import timezone

from apps.src.models import Account, Tizorbank

APILAYER_KEY = settings.APILAYER_KEY

class AddFunds(CronJobBase):
    RUN_EVERY_MINS = 1

    schedule = Schedule(run_every_mins=RUN_EVERY_MINS)
    code = 'apps.src.AddFunds'
    
    #crontab -e
    #* * * * * /apps/Tizorbank/venv/bin/python /apps/Tizorbank/core/manage.py runcrons
    def do(self):
        """cron Configuration
        Install django-crontab and add 'django_cron' to INSTALLED_APPS
        The `cronjobs` list needs to add config to settings.py:
            CRON_CLASSES = [
                "apps.core.cron.AddFunds",
            ]
            
        before add cronjobs clases add the crontask to crontab
        $ crontab -e
        > 0 */12 * * * /app/Tizorbank/core/venv/bin/python /app/Tizorbank/core/manage.py runcrons
        """

        setting = Tizorbank.objects.get(default="Tizorbank")
        eDate = timezone.now().strftime("%Y-%m-%d %H:%M")

        try:
            url = "https://api.coingecko.com/api/v3/simple/price?ids=bitcoin&vs_currencies=usd"
            response = requests.get(url)
            if response.status_code == 200:
                data = response.json()
                latestBTC=data["bitcoin"]["usd"]
                setting.latestBTC = latestBTC
                setting.save()

        except Exception as e:
            with open(os.path.join(settings.BASE_DIR, 'logs/core.log'), 'a') as f:
                f.write("CronBTC {} --> Error: {}\n".format(eDate, str(e)))

        try:
            url = "https://api.apilayer.com/fixer/latest?base=USD&symbols=COP"
            headers = {
                'apikey': f'{APILAYER_KEY}'
            }
            response = requests.get(url, headers=headers)
            if response.status_code == 200:
                data = response.json()
                latestUSD=data["rates"]["COP"]
                setting.latestUSD = latestUSD
                setting.save()

        except Exception as e:
            
            with open(os.path.join(settings.BASE_DIR, 'logs/core.log'), 'a') as f:
                f.write("CronUSD {} --> Error: {}\n".format(eDate, str(e)))    
        
        with open(os.path.join(settings.BASE_DIR, 'logs/logcron.txt'), 'a') as f:
            f.write("ServiceCron Active {}\n".format(eDate))

        accounts = Account.objects.all()

        for obj in accounts:
            
            dayli_interest = float((obj.interest)/(100*30))
            dayli_ammount = int(obj.ammount*(dayli_interest))
            available = int(obj.total-obj.paid + dayli_ammount)
            obj.available = available
            obj.save()
                        
            file_patch = os.path.join(settings.BASE_DIR, 'logs/users/'+ obj.username + '.xlsx')
            
            try:
                if not os.path.exists(file_patch):
                    WB = Workbook()
                    WS = WB.active
                    WS.append(["type","date","$daily","$aviable","$withdrawal","$invoice","$total"])
                else:
                    WB = load_workbook(file_patch)
                    WS = WB.active

                data = [1, timezone.now().strftime("%Y-%m-%d"), dayli_ammount, obj.available, 0, 0, obj.total]

                WS.append(data)
                WB.save(file_patch)
                
            except Exception as e:
                with open(os.path.join(settings.BASE_DIR, 'logs/workbook.log'), 'a') as f:
                    f.write("CronJob WorkbookError: {}\n".format(str(e)))

cronjobs = [
    AddFunds,
]



